# -*- coding: utf-8 -*-
import os, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# =========================
# Utils
# =========================
def strip_tz(dt):
    """Excel ne supporte pas les datetime avec timezone"""
    return dt.replace(tzinfo=None) if hasattr(dt, "tzinfo") and dt.tzinfo else dt

def clean_sheet_title(title: str) -> str:
    cleaned = re.sub(r'[:\\/*?\[\]]', '_', str(title))
    return cleaned[:25]

# =========================
# Fonction build_excel générique
# =========================
def build_excel(df, bookmaker_name=None, export_dir=".", kelly_number=4, stake_number=15):
    # Si aucun nom de bookmaker fourni, on prend le premier présent dans df
    if bookmaker_name is None:
        if "Bookmaker" in df.columns and not df["Bookmaker"].empty:
            bookmaker_name = df["Bookmaker"].iloc[0]
        else:
            bookmaker_name = "Bookmaker"  # fallback
            
    # Styles communs
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Headers Excel
    headers = [
        "Extraction",
        f"Cutoff_{bookmaker_name}",
        "Competition",
        "Evenement",
        f"Competiteur_{bookmaker_name}",
        f"Cote_{bookmaker_name}",
        "Cote_PS3838",      # G
        "TrueOdds_MPTO",    # H
        "ImpliedProb",      # I
        "TrueProb_MPTO",    # J
        "TRJ",              # K
        "%_boost",          # L
        f"Kelly_{kelly_number}",   # M
        f"Stake_{stake_number}",   # N
        "Potential_Payout",          # O
        "Surebet",                 # P
        "TRJ_Book"                 # Q
            ]
    
    
    # =========================
    # Résumé des compétitions
    # =========================
    cutoff_col = f"Cutoff_{bookmaker_name}" if f"Cutoff_{bookmaker_name}" in df.columns else "Cutoff"
    competiteur_col = f"Competiteur_{bookmaker_name}" if f"Competiteur_{bookmaker_name}" in df.columns else "Competiteur"
    
    # On prend l'index du cutoff le plus récent par compétition
    idx = df.groupby("Competition")[cutoff_col].idxmax()
    summary = df.loc[idx].copy()
    
    # Ajouter le nombre de cotes total par compétition
    summary["Nb_Cotes"] = df.groupby("Competition")[competiteur_col].transform("count").loc[idx]
    
    # Trier par cutoff le plus proche (celui qui arrivera en premier)
    summary = summary.sort_values(cutoff_col)
    
    # Affichage console lisible avec saut de ligne
    print("\n📊 Résumé des compétitions:\n")
    for _, row in summary.iterrows():
        print(f"\n- {row['Competition']} | Cutoff: {row[cutoff_col]} | Nb Cotes: {row['Nb_Cotes']}\n")


    # Création du workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Sélection de l’index du cutoff le plus récent par compétition
    cutoff_col = f"Cutoff_{bookmaker_name}" if f"Cutoff_{bookmaker_name}" in df.columns else "Cutoff"
    idx = df.groupby("Competition")[cutoff_col].idxmax()
    
    # DataFrame regroupé par compétition et trié selon le cutoff le plus proche
    df_comp_sorted = df.loc[idx].sort_values(cutoff_col)
    competitions_sorted = df_comp_sorted["Competition"].tolist()
    
    # Boucle sur les compétitions triées
    for competition in competitions_sorted:
        df_group = df[df["Competition"] == competition]
        sheet_title = clean_sheet_title(competition)
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_view.zoomScale = 70

        # Ajouter headers
        ws.append(headers)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Remplissage des lignes
        for idx, row in enumerate(df_group.itertuples(index=False), start=2):
            stake_value = f'VALUE(RIGHT($N$1,LEN($N$1)-FIND("_",$N$1)))'
            kelly_fraction = f'VALUE(RIGHT($M$1,LEN($M$1)-FIND("_",$M$1)))'

            ws.append([
                strip_tz(getattr(row, "Extraction", None)),
                strip_tz(getattr(row, f"Cutoff_{bookmaker_name}", getattr(row, "Cutoff", None))),
                getattr(row, "Competition", None),
                getattr(row, "Evenement", None),
                getattr(row, f"Competiteur_{bookmaker_name}", getattr(row, "Competiteur", None)),
                getattr(row, f"Cote_{bookmaker_name}", getattr(row, "Cote", None)),
                "",  # G
                f'=IF(G{idx}<>"",( (COUNTIF(D:D,D{idx})*G{idx}) / (COUNTIF(D:D,D{idx}) - (SUMIF(D:D,D{idx},I:I)-1)*G{idx}) ),"")',  # H
                f'=IF(G{idx}<>"",1/G{idx},"")',    # I
                f'=IF(H{idx}<>"",1/H{idx},"")',    # J
                "",                                # K
                f'=IF(AND(F{idx}<>"",H{idx}<>""),F{idx}/H{idx}-1,"")', # L 
                f'=IF(H{idx}="","",((F{idx}-1)*(1/H{idx})-(1-(1/H{idx})))/(F{idx}-1)/{kelly_fraction})', # M 
                f'=IFERROR(IF(H{idx}="","",((F{idx}-1)*(1/H{idx})-(1-(1/H{idx})))/(F{idx}-1)/{kelly_fraction}*{stake_value})*100,"")', # N              
                f'=IF(OR(F{idx}="",N{idx}=""), "", F{idx}*N{idx})',  # O Potential_Payout
                "",  # P Surebet
                ""   # Q TRJ_Book
            ])

        # TRJ et Surebet identique à avant
        for i in range(0, len(df_group), 2):
            idx1 = i + 2
            idx2 = i + 3
            if idx2 <= ws.max_row:
                ws.cell(row=idx1, column=11).value = f'=IF(AND(F{idx1}<>"",G{idx2}<>""),1/((1/F{idx1})+(1/G{idx2})),"")'
                ws.cell(row=idx2, column=11).value = f'=IF(AND(F{idx2}<>"",G{idx1}<>""),1/((1/F{idx2})+(1/G{idx1})),"")'
                ws.cell(row=idx1, column=16).value = f'=IF(AND(F{idx1}<>"",G{idx2}<>""),IF(1/((1/F{idx1})+(1/G{idx2}))>1,"YES","NO"),"")'
                ws.cell(row=idx2, column=16).value = f'=IF(AND(F{idx2}<>"",G{idx1}<>""),IF(1/((1/F{idx2})+(1/G{idx1}))>1,"YES","NO"),"")'
                ws.cell(row=idx1, column=17).value = f'=IF(AND(F{idx1}<>"",F{idx2}<>""),1/((1/F{idx1})+(1/F{idx2})),"")'
                ws.cell(row=idx2, column=17).value = f'=IF(AND(F{idx2}<>"",F{idx1}<>""),1/((1/F{idx2})+(1/F{idx1})),"")'

        # Mise en forme identique à avant
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row_cells:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        for col in ws.iter_cols(min_col=1, max_col=6, max_row=ws.max_row):
            max_length = 0
            column = col[0].column
            column_letter = get_column_letter(column)
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max(12, max_length + 0.25)

        for col in ws.iter_cols(min_col=7, max_col=ws.max_column):
            column_letter = get_column_letter(col[0].column)
            ws.column_dimensions[column_letter].width = 14

        for ridx in range(2, ws.max_row + 1):
            ws[f"H{ridx}"].number_format = '0.000'
            ws[f"I{ridx}"].number_format = '0.000'
            ws[f"J{ridx}"].number_format = '0.000'
            ws[f"K{ridx}"].number_format = '0.0%'
            ws[f"L{ridx}"].number_format = '0.0%'
            ws[f"M{ridx}"].number_format = '0.0%'
            ws[f"N{ridx}"].number_format = '€#,##0'
            ws[f"O{ridx}"].number_format = '€#,##0'

        # Conditional formatting identique à avant
        last_col_letter = get_column_letter(ws.max_column)
        
        ws.conditional_formatting.add(f"A2:{last_col_letter}{ws.max_row}",FormulaRule(formula=[f'$P2=""'], fill=white_fill))
        ws.conditional_formatting.add(f"A2:{last_col_letter}{ws.max_row}",FormulaRule(formula=[f'AND($P2="NO",$L2<0)'], fill=white_fill))
        ws.conditional_formatting.add(f"A2:{last_col_letter}{ws.max_row}",FormulaRule(formula=[f'AND($L2>0,$P2<>"YES")'], fill=orange_fill))
        ws.conditional_formatting.add(f"A2:{last_col_letter}{ws.max_row}",FormulaRule(formula=[f'$P2="YES"'], fill=green_fill)) 



    os.makedirs(export_dir, exist_ok=True)
    date_str = datetime.today().strftime("%Y-%m-%d")
    filename = f"Extract_{bookmaker_name}_{date_str}.xlsx"
    full_path = os.path.join(export_dir, filename)
    wb.save(full_path)
    return full_path

